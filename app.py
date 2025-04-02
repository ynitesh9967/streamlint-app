import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook

def process_script_1(file):
    df = pd.read_excel(file, dtype=str)
    df = df.map(lambda x: x.replace("'", "") if isinstance(x, str) else x)
    df.columns = df.columns.str.strip()
    df["REC FMT"] = df["REC FMT"].str.strip()
    
    if "DOMESTIC AMT" in df.columns:
        df["DOMESTIC AMT"] = pd.to_numeric(df["DOMESTIC AMT"], errors="coerce")
    
    df_bat = df[df["REC FMT"] == "BAT"]
    df_cvd = df[df["REC FMT"] == "CVD"]
    
    df_bat["Vlookup with CVD"] = df_bat["MERCHANT_TRACKID"].map(df_cvd.set_index("MERCHANT_TRACKID").index.to_series())
    df_cvd["Vlookup with BAT"] = df_cvd["MERCHANT_TRACKID"].map(df_bat.set_index("MERCHANT_TRACKID").index.to_series())
    
    if "SETTLE DATE" in df_cvd.columns:
        df_cvd["SETTLE DAY"] = pd.to_datetime(df_cvd["SETTLE DATE"], errors="coerce").dt.day.astype("Int64")
    if "SETTLE DATE" in df_bat.columns:
        df_bat["SETTLE DAY"] = pd.to_datetime(df_bat["SETTLE DATE"], errors="coerce").dt.day.astype("Int64")
    
    df_cvd["Remark"] = df_cvd.apply(lambda row: "Auto Reversal" if pd.notna(row["Vlookup with BAT"]) else f"Pending Refund {row['SETTLE DAY']}" if pd.notna(row["SETTLE DAY"]) else "Pending Refund", axis=1)
    df_bat["Remark"] = df_bat.apply(lambda row: "Auto Reversal" if pd.notna(row["Vlookup with CVD"]) else f"Pending Refund {row['SETTLE DAY']}" if pd.notna(row["SETTLE DAY"]) else "Pending Refund", axis=1)
    
    bat_pending = df_bat[df_bat["Remark"].str.startswith("Pending", na=False)].copy()
    bat_pending["UDF1"] = "'" + bat_pending["UDF1"].fillna('') + "',"
    bat_pending_final = bat_pending[["Remark", "UDF1"]]
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_cvd.to_excel(writer, sheet_name='CVD', index=False)
        df_bat.to_excel(writer, sheet_name='BAT', index=False)
        bat_pending_final.to_excel(writer, sheet_name='Pending Refunds', index=False)
    output.seek(0)
    return output

from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_duplicates(ws, df, key_column):

  duplicate_rows = df[df.duplicated(subset=[key_column], keep=False)].index + 2  # Excel row index starts at 1

  red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

  for row in duplicate_rows:
    for col in range(1, ws.max_column + 1):
      ws.cell(row=row, column=col).fill = red_fill

def process_script_2(file2, file1_output):
  
  df3 = pd.read_excel(file2, dtype=str)
  df3.columns = df3.columns.str.strip()
  df3["transaction_uid"] = df3["transaction_uid"].astype(str).str.strip()
  df3["state"] = df3["state"].astype(str).str.strip()
  
  xls = pd.ExcelFile(file1_output)
  bat_df = xls.parse("BAT")
  cvd_df = xls.parse("CVD")
    
  bat_df.columns = bat_df.columns.str.strip()
  bat_df["UDF1"] = bat_df["UDF1"].astype(str).str.strip()

  bat_df = bat_df.merge(
      df3[["transaction_uid", "state"]],
      left_on="UDF1",
      right_on="transaction_uid",
      how="left"
    )

    
  bat_df.drop(columns=["transaction_uid"], inplace=True)
  

  bat_summary = bat_df.groupby("SETTLE DATE").agg(
      Bank_MPR_Count=("DOMESTIC AMT", "count"),
      Bank_MPR_Amount=("DOMESTIC AMT", "sum")
  ).reset_index()

  cvd_summary = cvd_df.groupby("SETTLE DATE").agg(
      Bank_MPR_Refund_Count=("DOMESTIC AMT", "count"),
      Bank_MPR_Refund_Amount=("DOMESTIC AMT", "sum")
  ).reset_index()

  pending_bat = bat_df[bat_df["Remark"].str.startswith("Pending", na=False)]
  pending_summary = pending_bat.groupby("SETTLE DATE").agg(
      Pending_Refund_Count=("DOMESTIC AMT", "count"),
      Pending_Refund_Sum=("DOMESTIC AMT", "sum")
  ).reset_index()

  summary_df = bat_summary.merge(cvd_summary, on="SETTLE DATE", how="outer")\
                      .merge(pending_summary, on="SETTLE DATE", how="outer")

  output = io.BytesIO()
  with pd.ExcelWriter(output, engine='openpyxl') as writer:
      bat_df.to_excel(writer, sheet_name="BAT", index=False)
      cvd_df.to_excel(writer, sheet_name="CVD", index=False)
      summary_df.to_excel(writer, sheet_name="summary_output", index=False)
  
  output.seek(0)
    
    # Reload workbook to apply formatting
  wb = load_workbook(output)
  ws_bat = wb["BAT"]
  ws_cvd = wb["CVD"]

  # Apply duplicate highlighting
  highlight_duplicates(ws_bat, bat_df, "MERCHANT_TRACKID")
  highlight_duplicates(ws_cvd, cvd_df, "MERCHANT_TRACKID")

  # Save workbook again
  new_output = io.BytesIO()
  wb.save(new_output)
  new_output.seek(0)
    
  return new_output

def process_script_3(yesterday_file, today_file):
    # --- Function to update remarks based on matching ---
    def update_sheet(df_yesterday, df_today, key_column):
        # Clean columns (exactly as in your logic)
        df_yesterday.columns = df_yesterday.columns.str.strip().str.upper()
        df_today.columns = df_today.columns.str.strip().str.upper()

        # Clean key column (your exact approach)
        df_yesterday[key_column] = df_yesterday[key_column].astype(str).str.strip()
        df_today[key_column] = df_today[key_column].astype(str).str.strip()

        # Initialize REMARK if missing (as in your code)
        if "REMARK" not in df_yesterday.columns:
            df_yesterday["REMARK"] = "Pending"
        else:
            df_yesterday["REMARK"] = df_yesterday["REMARK"].astype(str).str.strip()

        # Your exact matching logic
        matching_values = df_yesterday[key_column].isin(df_today[key_column])
        pending_mask = df_yesterday["REMARK"].str.lower().str.startswith('pending', na=False)
        df_yesterday.loc[matching_values & pending_mask, "REMARK"] = "System Refund"

        return df_yesterday

    # Load files into dataframes (following your approach)
    df_yesterday_bat = pd.read_excel(yesterday_file, sheet_name="BAT", dtype=str)
    df_today_bat = pd.read_excel(today_file, sheet_name="BAT", dtype=str)
    df_yesterday_cvd = pd.read_excel(yesterday_file, sheet_name="CVD", dtype=str)
    df_today_cvd = pd.read_excel(today_file, sheet_name="CVD", dtype=str)

    # Process BAT and CVD sheets with your key columns
    updated_bat = update_sheet(df_yesterday_bat, df_today_bat, "VLOOKUP WITH CVD")
    updated_cvd = update_sheet(df_yesterday_cvd, df_today_cvd, "VLOOKUP WITH BAT")

    # Your exact column validation
    required_cols_bat = ["SETTLE DATE", "DOMESTIC AMT"]
    required_cols_cvd = ["SETTLE DATE", "DOMESTIC AMT", "REMARK"]
    
    if not all(col in updated_bat.columns for col in required_cols_bat):
        st.error("Missing required columns in 'BAT' sheet")
        return None
    if not all(col in updated_cvd.columns for col in required_cols_cvd):
        st.error("Missing required columns in 'CVD' sheet")
        return None

    # Convert amounts (your exact approach)
    updated_bat['DOMESTIC AMT'] = pd.to_numeric(updated_bat['DOMESTIC AMT'], errors='coerce')
    updated_cvd['DOMESTIC AMT'] = pd.to_numeric(updated_cvd['DOMESTIC AMT'], errors='coerce')

    # Your exact summary calculation logic
    bat_summary = updated_bat.groupby("SETTLE DATE").agg(
        Bank_MPR_Count=("DOMESTIC AMT", "count"),
        Bank_MPR_Amount=("DOMESTIC AMT", "sum")
    ).reset_index()

    cvd_summary = updated_cvd.groupby("SETTLE DATE").agg(
        Bank_MPR_Refund_Count=("DOMESTIC AMT", "count"),
        Bank_MPR_Refund_Amount=("DOMESTIC AMT", "sum")
    ).reset_index()

    pending_cvd = updated_cvd[updated_cvd["REMARK"].str.lower().str.startswith("pending", na=False)]
    pending_summary = pending_cvd.groupby("SETTLE DATE").agg(
        Pending_Refund_Count=("DOMESTIC AMT", "count"),
        Pending_Refund_Sum=("DOMESTIC AMT", "sum")
    ).reset_index()

    # Your exact merge logic
    summary_df = bat_summary.merge(cvd_summary, on="SETTLE DATE", how="outer")\
                           .merge(pending_summary, on="SETTLE DATE", how="outer")

    # Create output (as in your code)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        updated_bat.to_excel(writer, sheet_name='BAT', index=False)
        updated_cvd.to_excel(writer, sheet_name='CVD', index=False)
        summary_df.to_excel(writer, sheet_name='summary_output', index=False)
    output.seek(0)
    
    return output
def main():
  st.title("Excel Processing Pipeline")
  step = st.sidebar.radio("Choose Step", ("Step 1", "Step 2", "Step 3"))

  if step == "Step 1":
    file = st.sidebar.file_uploader("Upload an Excel File", type=["xls", "xlsx"], key="input_file")

    if file:

      processed_file = process_script_1(file)
      st.download_button("Download Processed File", processed_file, file_name="processed_data.xlsx")

  elif step == "Step 2":
    file2 = st.sidebar.file_uploader("Upload Query Excel File", type=["xls", "xlsx"], key="input_file2")
    file1_output = st.sidebar.file_uploader("Upload Processed Data (BAT & CVD)", type=["xls", "xlsx"], key="input_file1_output")
    if file2 and file1_output:
      updated_processed_file = process_script_2(file2, file1_output)
      st.download_button("Download Updated Processed File", updated_processed_file, file_name="updated_processed_data.xlsx")

  elif step == "Step 3":
    st.title("Excel Processing Pipeline - Step 3")
    yesterday_file = st.sidebar.file_uploader("Upload Yesterday's Data", type=["xls", "xlsx"], key="input_file_yesterday")
    today_file = st.sidebar.file_uploader("Upload Today's Data", type=["xls", "xlsx"], key="input_file_today")
    
    if yesterday_file and today_file:
        final_output = process_script_3(yesterday_file, today_file)
        if final_output:
            st.download_button("Download Updated Yesterday File", final_output, file_name="updated_yesterday.xlsx")

if __name__ == "__main__":
    main()

